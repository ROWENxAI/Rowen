"""
全量知识资产处理脚本 - 输出到 H:\转录
支持 D:\转录 + G:\转录 全部文件
Obsidian 双链格式 + 详细内容
"""
import os, sys, json, time, subprocess, traceback
from datetime import datetime
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

OUTPUT_ROOT = r"H:\转录"
LOG_DIR = os.path.join(OUTPUT_ROOT, "任务日志")
os.makedirs(LOG_DIR, exist_ok=True)

DONE_LOG = os.path.join(LOG_DIR, "processed_files.json")
ERROR_LOG = os.path.join(LOG_DIR, "errors.log")
PROGRESS_LOG = os.path.join(LOG_DIR, "progress.log")

MEDIA_EXTS = {'.mp4', '.mp3', '.mov', '.avi', '.mkv', '.m4a', '.wav'}
DOC_EXTS = {'.pdf', '.epub', '.docx', '.doc', '.txt', '.xlsx'}
SKIP_DIR_NAMES = {'任务日志', 'X平台知识库', '.browser_data', '__pycache__', '.git'}

def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    try:
        print(line, flush=True)
    except UnicodeEncodeError:
        print(line.encode('utf-8', errors='replace').decode('utf-8'), flush=True)
    with open(PROGRESS_LOG, "a", encoding="utf-8") as f:
        f.write(line + "\n")

def log_error(filepath, reason):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(ERROR_LOG, "a", encoding="utf-8") as f:
        f.write(f"[{ts}] {filepath}: {reason}\n")

def load_done():
    if os.path.exists(DONE_LOG):
        with open(DONE_LOG, "r", encoding="utf-8") as f:
            return set(json.load(f))
    return set()

def save_done(done_set):
    with open(DONE_LOG, "w", encoding="utf-8") as f:
        json.dump(list(done_set), f, ensure_ascii=False, indent=2)

def get_duration(filepath):
    try:
        result = subprocess.run(
            ["ffprobe", "-v", "quiet", "-show_entries", "format=duration",
             "-of", "default=noprint_wrappers=1:nokey=1", filepath],
            capture_output=True, text=True, timeout=30
        )
        return float(result.stdout.strip())
    except:
        return 0

def format_ts(seconds):
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = int(seconds % 60)
    if h > 0:
        return f"{h:02d}:{m:02d}:{s:02d}"
    return f"{m:02d}:{s:02d}"

def safe_filename(name):
    for ch in ['<', '>', ':', '"', '/', '\\', '|', '?', '*']:
        name = name.replace(ch, '_')
    return name[:200]

def get_output_path(src_path, src_root, output_root):
    """Map source path to output path maintaining directory structure."""
    rel = os.path.relpath(src_path, src_root)
    return os.path.join(output_root, rel)

def extract_audio(video_path, audio_path):
    try:
        subprocess.run(
            ["ffmpeg", "-i", video_path, "-vn", "-acodec", "pcm_s16le",
             "-ar", "16000", "-ac", "1", "-y", audio_path],
            capture_output=True, timeout=600
        )
        return os.path.exists(audio_path)
    except:
        return False

def transcribe_whisper(filepath, model, is_audio):
    temp_audio = None
    try:
        if is_audio:
            audio_path = filepath
        else:
            temp_audio = os.path.join(os.environ.get("TEMP", r"C:\temp"),
                                       f"whisper_{os.getpid()}.wav")
            os.makedirs(os.path.dirname(temp_audio), exist_ok=True)
            if not extract_audio(filepath, temp_audio):
                return None, "Failed to extract audio"
            audio_path = temp_audio

        segments, info = model.transcribe(
            audio_path, language="zh", beam_size=5,
            vad_filter=True, vad_parameters=dict(min_silence_duration_ms=500)
        )
        result = []
        for seg in segments:
            result.append({"start": round(seg.start, 2), "end": round(seg.end, 2), "text": seg.text.strip()})
        return result, info
    except Exception as e:
        return None, str(e)
    finally:
        if temp_audio and os.path.exists(temp_audio):
            try: os.remove(temp_audio)
            except: pass

def write_obsidian_transcript(segments, filename, duration, output_path, source_path):
    """Write full transcript in Obsidian format with detailed content."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    tag = "转录/逐字稿"
    md = f"---\ntitle: \"{filename}\"\ntype: transcript\nduration: \"{format_ts(duration)}\"\n"
    md += f"date: {datetime.now().strftime('%Y-%m-%d')}\ntags: [{tag}]\n"
    md += f"source: \"{source_path}\"\n---\n\n"
    md += f"# {filename} 逐字稿\n\n"
    md += f"> 时长: {format_ts(duration)} | 转录时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"

    full_text = ""
    for seg in segments:
        ts = format_ts(seg["start"])
        md += f"**`[{ts}]`** {seg['text']}\n\n"
        full_text += seg["text"] + " "

    md += f"\n---\n\n> [!note] 全文统计\n> 总字数: {len(full_text)} | 段落数: {len(segments)}\n"
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(md)
    return full_text

def write_obsidian_study_notes(segments, filename, duration, output_path, source_path, full_text):
    """Write detailed study notes in Obsidian format."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    tag = "转录/学习笔记"

    md = f"---\ntitle: \"{filename} 学习笔记\"\ntype: study-notes\nduration: \"{format_ts(duration)}\"\n"
    md += f"date: {datetime.now().strftime('%Y-%m-%d')}\ntags: [{tag}]\n"
    md += f"source: \"{source_path}\"\n---\n\n"
    md += f"# {filename} 学习笔记\n\n"
    md += f"> 时长: {format_ts(duration)} | 来源: `{os.path.basename(source_path)}`\n\n"

    md += "## 内容概要\n\n"
    section_dur = max(300, duration / 10)
    sections = []
    cur = []
    cur_start = 0
    for seg in segments:
        if seg["start"] >= cur_start + section_dur and cur:
            sections.append((cur_start, cur))
            cur_start = seg["start"]
            cur = []
        cur.append(seg)
    if cur:
        sections.append((cur_start, cur))

    for idx, (start, segs) in enumerate(sections, 1):
        text = " ".join(s["text"] for s in segs)
        end = segs[-1]["end"]
        md += f"### 第{idx}部分 `[{format_ts(start)}-{format_ts(end)}]`\n\n"
        md += text[:800] + "\n\n"

    md += "## 关键词提取\n\n"
    words = full_text.split()
    md += f"- 总词数: ~{len(words)}\n"
    md += f"- 总字数: {len(full_text)}\n\n"

    md += "## 完整内容\n\n"
    md += f"详见 [[{os.path.basename(output_path).replace('_学习笔记', '_逐字稿')}|逐字稿]]\n\n"

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(md)

def extract_pdf_text(filepath):
    try:
        import fitz
        doc = fitz.open(filepath)
        text = ""
        for page in doc:
            text += page.get_text()
        doc.close()
        return text
    except:
        pass
    try:
        from pdfminer.high_level import extract_text
        return extract_text(filepath)
    except:
        pass
    return None

def extract_epub_text(filepath):
    try:
        import ebooklib
        from ebooklib import epub
        from bs4 import BeautifulSoup
        book = epub.read_epub(filepath)
        text = ""
        for item in book.get_items():
            if item.get_type() == 9:
                soup = BeautifulSoup(item.get_content(), 'html.parser')
                text += soup.get_text() + "\n"
        return text
    except:
        return None

def extract_docx_text(filepath):
    try:
        from docx import Document
        doc = Document(filepath)
        return "\n".join(p.text for p in doc.paragraphs)
    except:
        return None

def extract_txt_text(filepath):
    for enc in ['utf-8', 'gbk', 'gb2312', 'latin-1']:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                return f.read()
        except:
            continue
    return None

def extract_xlsx_text(filepath):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        text = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            text += f"\n=== Sheet: {sheet} ===\n"
            for row in ws.iter_rows(values_only=True):
                text += "\t".join(str(c) if c is not None else "" for c in row) + "\n"
        wb.close()
        return text
    except:
        return None

def write_obsidian_doc_notes(text, filename, output_path, source_path):
    """Write document knowledge notes in Obsidian format."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    tag = "文档/知识梳理"

    md = f"---\ntitle: \"{filename} 知识梳理\"\ntype: doc-notes\ndate: {datetime.now().strftime('%Y-%m-%d')}\n"
    md += f"tags: [{tag}]\nsource: \"{source_path}\"\n---\n\n"
    md += f"# {filename} 知识梳理\n\n"
    md += f"> 来源: `{os.path.basename(source_path)}` | 字数: {len(text)}\n\n"

    paras = [p.strip() for p in text.split('\n') if p.strip()]
    md += "## 内容全文\n\n"
    for p in paras[:500]:
        md += p + "\n\n"
    if len(paras) > 500:
        md += f"\n> [!tip] 共 {len(paras)} 段，以上显示前 500 段\n"

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(md)

def process_media(filepath, src_root, model, done):
    filename = os.path.basename(filepath)
    name_no_ext = os.path.splitext(filename)[0]
    ext = os.path.splitext(filename)[1].lower()
    is_audio = ext in {'.mp3', '.m4a', '.wav'}

    if is_audio:
        out_dir = os.path.join(OUTPUT_ROOT, "音频转录")
    else:
        out_dir = os.path.join(OUTPUT_ROOT, "视频转录")

    rel = os.path.relpath(os.path.dirname(filepath), src_root)
    out_subdir = os.path.join(out_dir, rel)
    safe_name = safe_filename(name_no_ext)

    transcript_path = os.path.join(out_subdir, f"{safe_name}_逐字稿.md")
    notes_path = os.path.join(out_subdir, f"{safe_name}_学习笔记.md")

    if os.path.exists(transcript_path) and os.path.exists(notes_path):
        done.add(filepath)
        save_done(done)
        return True

    duration = get_duration(filepath)
    if duration == 0:
        log_error(filepath, "Cannot get duration")
        return False

    t0 = time.time()
    segments, info = transcribe_whisper(filepath, model, is_audio)
    t1 = time.time()

    if segments is None:
        log_error(filepath, str(info))
        return False

    full_text = write_obsidian_transcript(segments, filename, duration, transcript_path, filepath)
    write_obsidian_study_notes(segments, filename, duration, notes_path, filepath, full_text)

    done.add(filepath)
    save_done(done)
    log(f"  {len(segments)} segments, {len(full_text)} chars, {t1-t0:.1f}s")
    return True

def process_doc(filepath, src_root, done):
    filename = os.path.basename(filepath)
    name_no_ext = os.path.splitext(filename)[0]
    ext = os.path.splitext(filename)[1].lower()

    out_dir = os.path.join(OUTPUT_ROOT, "文档笔记")
    rel = os.path.relpath(os.path.dirname(filepath), src_root)
    out_subdir = os.path.join(out_dir, rel)
    safe_name = safe_filename(name_no_ext)

    notes_path = os.path.join(out_subdir, f"{safe_name}_知识梳理.md")
    if os.path.exists(notes_path):
        done.add(filepath)
        save_done(done)
        return True

    extractors = {
        '.pdf': extract_pdf_text, '.epub': extract_epub_text,
        '.docx': extract_docx_text, '.doc': extract_docx_text,
        '.txt': extract_txt_text, '.xlsx': extract_xlsx_text,
    }
    extractor = extractors.get(ext)
    if not extractor:
        log_error(filepath, f"No extractor for {ext}")
        return False

    text = extractor(filepath)
    if not text or len(text.strip()) < 20:
        log_error(filepath, "No text extracted")
        done.add(filepath)
        save_done(done)
        return False

    write_obsidian_doc_notes(text, filename, notes_path, filepath)
    done.add(filepath)
    save_done(done)
    log(f"  {len(text)} chars extracted")
    return True

def scan_files(root_dir):
    media_files = []
    doc_files = []
    # Skip wav from G drive (test files)
    is_g_drive = root_dir.upper().startswith("G:")
    media_filter = MEDIA_EXTS - ({'.wav'} if is_g_drive else set())

    # For G drive, only process course directories (with mp4/mp3)
    # For D drive, process everything
    code_skip_exts = {'.py', '.pyc', '.js', '.php', '.h', '.css', '.html', '.json',
                      '.log', '.dat', '.dll', '.exe', '.db', '.ldb', '.db-journal',
                      '.old', '.hyb', '.fingerprint', '.pma', '.0-0', '.v2', '.bat',
                      '.sig', '.gif', '.ds_store', '.xlsx'}

    for dirpath, dirnames, filenames in os.walk(root_dir):
        dirnames[:] = [d for d in dirnames if d not in SKIP_DIR_NAMES]
        for f in filenames:
            ext = os.path.splitext(f)[1].lower()
            full_path = os.path.join(dirpath, f)
            # Skip test/sample files
            if 'test' in f.lower() or '样本' in f or 'sample' in f.lower():
                continue
            if ext in media_filter:
                media_files.append(full_path)
            elif ext in DOC_EXTS and ext not in code_skip_exts:
                # For G drive, skip tiny txt files (likely code/config)
                if is_g_drive and ext == '.txt':
                    if os.path.getsize(full_path) < 10000:
                        continue
                doc_files.append(full_path)
    return media_files, doc_files

def main():
    log("=" * 60)
    log("=== Full Knowledge Processing -> H:\\转录 ===")
    log("=" * 60)

    # Scan both sources
    all_media = []
    all_docs = []

    for src_root in [r"D:\转录", r"G:\转录"]:
        if not os.path.exists(src_root):
            log(f"Source not found: {src_root}")
            continue
        log(f"Scanning {src_root}...")
        media, docs = scan_files(src_root)
        log(f"  Found {len(media)} media, {len(docs)} docs")
        all_media.extend((f, src_root) for f in media)
        all_docs.extend((f, src_root) for f in docs)

    all_media.sort(key=lambda x: os.path.getsize(x[0]))
    all_docs.sort(key=lambda x: os.path.getsize(x[0]))

    log(f"Total: {len(all_media)} media, {len(all_docs)} docs")

    done = load_done()
    remaining_media = [(f, r) for f, r in all_media if f not in done]
    remaining_docs = [(f, r) for f, r in all_docs if f not in done]
    log(f"Already done: {len(done)}, remaining media: {len(remaining_media)}, remaining docs: {len(remaining_docs)}")

    # Process docs first (fast)
    if remaining_docs:
        log(f"\n--- Processing {len(remaining_docs)} documents ---")
        for i, (fp, src_root) in enumerate(remaining_docs):
            try:
                fn = os.path.basename(fp)
                log(f"[Doc {i+1}/{len(remaining_docs)}] {fn}")
                process_doc(fp, src_root, done)
            except Exception as e:
                log(f"  ERROR: {e}")
                log_error(fp, traceback.format_exc())

    # Process media (slow, needs whisper)
    if remaining_media:
        log(f"\n--- Loading whisper model ---")
        from faster_whisper import WhisperModel
        model = WhisperModel("small", device="cuda", compute_type="float16")
        log("Model loaded!")

        log(f"\n--- Processing {len(remaining_media)} media files ---")
        start_time = time.time()
        for i, (fp, src_root) in enumerate(remaining_media):
            try:
                fn = os.path.basename(fp)
                log(f"[Media {i+1}/{len(remaining_media)}] {fn}")
                process_media(fp, src_root, model, done)

                elapsed = time.time() - start_time
                rate = (i + 1) / elapsed * 3600
                eta = (len(remaining_media) - i - 1) / (rate / 3600) if rate > 0 else 0
                if (i + 1) % 10 == 0:
                    log(f"  Progress: {i+1}/{len(remaining_media)} | ETA: {eta:.1f}h")
            except Exception as e:
                log(f"  ERROR: {e}")
                log_error(fp, traceback.format_exc())

    # Generate Obsidian index
    log("\n--- Generating Obsidian index ---")
    generate_obsidian_index()

    log("=" * 60)
    log(f"=== Processing complete: {len(done)} files ===")
    log("=" * 60)

def generate_obsidian_index():
    """Generate MOC (Map of Content) for Obsidian."""
    index_path = os.path.join(OUTPUT_ROOT, "00-知识库总览.md")
    md = "---\ntitle: 知识库总览\ntype: moc\ndate: "
    md += datetime.now().strftime('%Y-%m-%d') + "\ntags: [总览, MOC]\n---\n\n"
    md += "# 知识库总览\n\n"
    md += f"> 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"

    # Scan output directories
    for category in ["音频转录", "视频转录", "文档笔记"]:
        cat_dir = os.path.join(OUTPUT_ROOT, category)
        if not os.path.exists(cat_dir):
            continue
        transcripts = []
        for root, dirs, files in os.walk(cat_dir):
            for f in files:
                if f.endswith('.md'):
                    transcripts.append(os.path.join(root, f))

        md += f"## {category} ({len(transcripts)} 个)\n\n"
        for fp in sorted(transcripts)[:100]:
            rel = os.path.relpath(fp, OUTPUT_ROOT)
            name = os.path.splitext(os.path.basename(fp))[0]
            md += f"- [[{rel}|{name}]]\n"
        if len(transcripts) > 100:
            md += f"- ... 共 {len(transcripts)} 个文件\n"
        md += "\n"

    # X platform
    x_dir = os.path.join(OUTPUT_ROOT, "X平台知识库")
    if os.path.exists(x_dir):
        md += "## X平台内容\n\n"
        for f in sorted(os.listdir(x_dir)):
            if f.endswith('.md'):
                md += f"- [[X平台知识库/{f}|{os.path.splitext(f)[0]}]]\n"
        md += "\n"

    with open(index_path, "w", encoding="utf-8") as f:
        f.write(md)
    log(f"Index saved: {index_path}")

if __name__ == "__main__":
    main()
